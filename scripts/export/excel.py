from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment


class ExportDataExcel:
    @staticmethod
    def create_output_excel(rank_list, output_file):
        headers = [["Iteracija", "Broj retka", "PrP", "Indentifikacijska oznaka investicije", "OBJEKT/PLANSKA STAVKA", "OZNAKA DV", "IME DALEKOVODA (KANDIDATA ZA REVITALIZACIJU)", "Index"],
                   ["Iteracija", "Broj retka", "PrP", "Indentifikacijska oznaka investicije", "OBJEKT/PLANSKA STAVKA", "TVORNIČKI BROJ", "NAZIV TRANSFORMATORA", "Index"],
                   ["Iteracija", "Broj retka", "PrP", "Indentifikacijska oznaka investicije", "OBJEKT/PLANSKA STAVKA", "TVORNIČKI BROJ", "NAZIV TRANSFORMATORA", "Index"],
                   ["Iteracija", "Broj retka", "PrP", "Indentifikacijska oznaka investicije", "OBJEKT/PLANSKA STAVKA", "ŠIFRA LOKACIJE", "NAZIV TRANSFORMATORSKE STANICE", "Index"]]

        grey_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

        workbook = Workbook()
        worksheet = workbook.active

        # Write header
        worksheet.append(headers[rank_list[0].cc.value - 1])

        # Write Data
        for i, obj in enumerate(rank_list):
            worksheet.append([i + 1, obj.row_number, obj.cc.name, obj.id, obj.item, obj.code, obj.name, obj.index])

        for col in range(8):
            worksheet[1][col].fill = grey_fill
            worksheet[1][col].alignment = Alignment(horizontal='center', vertical='center')

        workbook.save(output_file)

from openpyxl import load_workbook
from scripts.invobj.investmentObject import InvestmentObject
from scripts.invobj.condition import Condition
from scripts.invobj.significance import Significance
from scripts.invobj.economy import Economy
from scripts.validation import Validate
from scripts.enumValues import ObjectType, ControlCenter


class ImportData:
    def __init__(self, main_window, file_path, object_type, observed_year):
        """
        :param main_window: main.MainWindow object, parent window
        :param file_path: string, input excel
        :param object_type: enumValues.ObjectType, type of investment object
        :param observed_year: int, reference year for caluclation
        """
        main_window.logDebug.info("Call:ImportData.__init__()")
        self.__pwin = main_window
        self.__observed_year = observed_year
        self.__max_age = 0
        self.__max_unavailability = 0
        self.__max_revitalization_infulance = 0
        self.__max_enpv = 0
        self.__max_transmitted_energy = 0
        self.__max_ratio = 0
        self.__max_energy = 0
        self.investmentObjects = []

        # Load workbook and worksheets
        try:
            self.__workbook = load_workbook(file_path, read_only=False, keep_vba=False, data_only=True, keep_links=False)
        except Exception as e:
            self.__pwin.error_message("Došlo je do pogreške prilikom učitavanja ulazne datoteke!")
            self.__pwin.logDebug.exception(e)
            raise Exception(e)
        try:
            self.__ws_condition = self.__workbook.get_sheet_by_name("I STANJE")
        except Exception as e:
            self.__pwin.error_message("Došlo je do pogreške prilikom učitavanja radnog lista 'I STANJE'")
            self.__pwin.logDebug.exception(e)
            raise Exception(e)
        try:
            self.__ws_significance = self.__workbook.get_sheet_by_name("II ZNAČAJ")
        except Exception as e:
            self.__pwin.error_message("Došlo je do pogreške prilikom učitavanja radnog lista 'II ZNAČAJ'")
            self.__pwin.logDebug.exception(e)
            raise Exception(e)
        try:
            if object_type != ObjectType.TransformerMedium:
                self.__ws_economy = self.__workbook.get_sheet_by_name("III EKONOMIJA")
            else:
                self.__ws_economy = None
        except Exception as e:
            self.__pwin.error_message("Došlo je do pogreške prilikom učitavanja radnog lista 'III EKONOMIJA'")
            self.__pwin.logDebug.exception(e)
            raise Exception(e)

        # Check headers
        err = self.check_headers(object_type)
        if err == 1:
            self.__pwin.error_message("Radni list 'I STANJE' nije ispravan!")
            self.__pwin.logDebug.exception("Worksheet 'I STANJE' has wrong header")
            raise Exception("Worksheet 'I STANJE' has wrong header")
        elif err == 2:
            self.__pwin.error_message("Radni list 'II ZNAČAJ' nije ispravan!")
            self.__pwin.logDebug.exception("Worksheet 'II ZNAČAJ' has wrong header")
            raise Exception("Worksheet 'II ZNAČAJ' has wrong header")
        elif err == 3:
            self.__pwin.error_message("Radni list 'III EKONOMIJA' nije ispravan!")
            self.__pwin.logDebug.exception("Worksheet 'III EKONOMIJA' has wrong header")
            raise Exception("Worksheet 'III EKONOMIJA' has wrong header")
        elif err == 4:
            raise Exception("Unknown object type")

        # Read data
        if object_type == ObjectType.Line:
            self.read_line()
        elif object_type == ObjectType.TransformerLarge:
            self.read_transformer_large()
        elif object_type == ObjectType.TransformerMedium:
            self.read_transformer_medium()
        elif object_type == ObjectType.Station:
            self.read_station()

        self.__pwin.logDebug.info("Return:ImportData.__init__()")

    # <editor-fold desc="#####  Check Headers  #####">
    def check_headers(self, object_type: ObjectType):
        if object_type == ObjectType.Line:
            if not self.check_condition_line():
                return 1
            elif not self.check_significance_line():
                return 2
            elif not self.check_economy_line():
                return 3
        elif object_type == ObjectType.TransformerLarge:
            if not self.check_condition_trafo_large():
                return 1
            elif not self.check_significance_trafo_large():
                return 2
            elif not self.check_economy_trafo_large():
                return 3
        elif object_type == ObjectType.TransformerMedium:
            if not self.check_condition_trafo_medium():
                return 1
            elif not self.check_significance_trafo_medium():
                return 2
            # elif not self.check_economy_trafo_medium():
            #     return 3
        elif object_type == ObjectType.Station:
            if not self.check_condition_station():
                return 1
            elif not self.check_significance_station():
                return 2
            elif not self.check_economy_station():
                return 3
        elif object_type != ObjectType.Line and object_type != ObjectType.TransformerLarge and \
                object_type != ObjectType.TransformerMedium and object_type != ObjectType.Station:
            return 4
        return 0
    # </editor-fold>

    # <editor-fold desc="#####  Check Condition  #####">
    def check_condition_line(self):
        if not all([self.__ws_condition[1][0].value == "PrP",
                    self.__ws_condition[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_condition[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_condition[1][3].value == "OZNAKA DV",
                    self.__ws_condition[1][4].value == "NAZIV DALEKOVODA",
                    self.__ws_condition[1][5].value == "S.1 Tehnička ispravnost jedinice",
                    self.__ws_condition[1][6].value == "(S.2) Starost jedinice",
                    self.__ws_condition[1][7].value == "(S.3) Neraspoloživost jedinice",
                    self.__ws_condition[1][8].value == "S.4 Rezultati pregleda i dijagnostike",
                    self.__ws_condition[1][9].value == "S.5 Ostali pokazatelji stanja"
                    ]):
            return False
        return True

    def check_condition_trafo_large(self):
        if not all([self.__ws_condition[1][0].value == "PrP",
                    self.__ws_condition[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_condition[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_condition[1][3].value == "TVORNIČKI BROJ",
                    self.__ws_condition[1][4].value == "NAZIV TRANSFORMATORA",
                    self.__ws_condition[1][5].value == "S.1 Rezultati pregleda i dijagnostike",
                    self.__ws_condition[1][6].value == "(S.2) Starost jedinice",
                    self.__ws_condition[1][7].value == "(S.3) Neraspoloživost"
                    ]):
            return False
        return True

    def check_condition_trafo_medium(self):
        if not all([self.__ws_condition[1][0].value == "PrP",
                    self.__ws_condition[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_condition[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_condition[1][3].value == "TVORNIČKI BROJ",
                    self.__ws_condition[1][4].value == "NAZIV TRANSFORMATORA",
                    self.__ws_condition[1][5].value == "S.1 Rezultati pregleda i dijagnostike",
                    self.__ws_condition[1][6].value == "(S.2) Starost jedinice",
                    self.__ws_condition[1][7].value == "(S.3) Neraspoloživost"
                    ]):
            return False
        return True

    def check_condition_station(self):
        if not all([self.__ws_condition[1][0].value == "PrP",
                    self.__ws_condition[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_condition[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_condition[1][3].value == "ŠIFRA LOKACIJE",
                    self.__ws_condition[1][4].value == "NAZIV TRANSFORMATORSKE STANICE",
                    self.__ws_condition[1][5].value == "(S.1) Starost",
                    self.__ws_condition[1][6].value == "S.2 Ocjena stanja primarne opreme",
                    self.__ws_condition[1][7].value == "S.3 Ocjena stanja sekundarne opreme",
                    self.__ws_condition[1][8].value == "S.4 Ocjena stanja pomoćne opreme",
                    self.__ws_condition[1][9].value == "S.5 Ocjena stanja građevinskih dijelova TS"
                    ]):
            return False
        return True
    # </editor-fold>

    # <editor-fold desc="#####  Check Significance  #####">
    def check_significance_line(self):
        if not all([self.__ws_significance[1][0].value == "PrP",
                    self.__ws_significance[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_significance[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_significance[1][3].value == "OZNAKA DV",
                    self.__ws_significance[1][4].value == "NAZIV DALEKOVODA",
                    self.__ws_significance[1][5].value == "Z.1 Poteba za povećanjem prijenosne moći voda",
                    self.__ws_significance[1][6].value == "Z.2 Utjecaj voda na sigurnost pogona prema kriteriju n-1",
                    self.__ws_significance[1][7].value == "Z.3 Rizik od trajne neraspoloživosti voda",
                    self.__ws_significance[1][8].value == "Z.4 Procjena šteta uzrokovanih neprovedbom projekta evitalizacije voda",
                    self.__ws_significance[1][9].value == "(Z.5) Utjecaj revitalizacije voda na gubitke snage u prijenosnoj mreži (MW)"
                    ]):
            return False
        return True

    def check_significance_trafo_large(self):
        if not all([self.__ws_significance[1][0].value == "PrP",
                    self.__ws_significance[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_significance[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_significance[1][3].value == "TVORNIČKI BROJ",
                    self.__ws_significance[1][4].value == "NAZIV TRANSFORMATORA",
                    self.__ws_significance[1][5].value == "Z.1 Sigurnost pogona",
                    self.__ws_significance[1][6].value == "Z.2 Štete",
                    self.__ws_significance[1][7].value == "Z.3 Rizik od trajne neraspoloživosti transformatora"
                    ]):
            return False
        return True

    def check_significance_trafo_medium(self):
        if not all([self.__ws_significance[1][0].value == "PrP",
                    self.__ws_significance[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_significance[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_significance[1][3].value == "TVORNIČKI BROJ",
                    self.__ws_significance[1][4].value == "NAZIV TRANSFORMATORA",
                    self.__ws_significance[1][5].value == "(Z.1) Prenesena energija (godišnja u GWh)",
                    self.__ws_significance[1][6].value == "(Z.2) Omjer između maksimalnog opterećenja i prividne snage",
                    self.__ws_significance[1][7].value == "Z.3 Štete"
                    ]):
            return False
        return True

    def check_significance_station(self):
        if not all([self.__ws_significance[1][0].value == "PrP",
                    self.__ws_significance[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_significance[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_significance[1][3].value == "ŠIFRA LOKACIJE",
                    self.__ws_significance[1][4].value == "NAZIV TRANSFORMATORSKE STANICE",
                    self.__ws_significance[1][5].value == "Z.1 Zadovoljenje opreme s obzirom na kratkospojnu razinu",
                    self.__ws_significance[1][6].value == "(Z.2) Godišnja energija kroz promatrano rasklopište",
                    self.__ws_significance[1][7].value == "Z.3 Rizik u slučaju zastoja jednog sabirničkog sustava",
                    self.__ws_significance[1][8].value == "Z.4 Štete u slučaju zastoja jednog sabirničkog sustava",
                    self.__ws_significance[1][9].value == "Z.5 Broj planiranih jedinica za priključenje na rasklopište"
                    ]):
            return False
        return True
    # </editor-fold>

    # <editor-fold desc="#####  Check Economy  #####">
    def check_economy_line(self):
        if not all([self.__ws_economy[1][0].value == "PrP",
                    self.__ws_economy[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_economy[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_economy[1][3].value == "OZNAKA DV",
                    self.__ws_economy[1][4].value == "NAZIV DALEKOVODA",
                    self.__ws_economy[1][5].value == "ENPV (HRK)"
                    ]):
            return False
        return True

    def check_economy_trafo_large(self):
        if not all([self.__ws_economy[1][0].value == "PrP",
                    self.__ws_economy[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_economy[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_economy[1][3].value == "TVORNIČKI BROJ",
                    self.__ws_economy[1][4].value == "NAZIV TRANSFORMATORA",
                    self.__ws_economy[1][5].value == "ENPV (HRK)"
                    ]):
            return False
        return True

    def check_economy_trafo_medium(self):
        if not all([self.__ws_economy[1][0].value == "PrP",
                    self.__ws_economy[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_economy[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_economy[1][3].value == "TVORNIČKI BROJ",
                    self.__ws_economy[1][4].value == "NAZIV TRANSFORMATORA",
                    self.__ws_economy[1][5].value == "ENPV (HRK)"
                    ]):
            return False
        return True

    def check_economy_station(self):
        if not all([self.__ws_economy[1][0].value == "PrP",
                    self.__ws_economy[1][1].value == "IDENTIFIKACIJSKA OZNAKA INVESTICIJE",
                    self.__ws_economy[1][2].value == "OBJEKT / PLANSKA STAVKA",
                    self.__ws_economy[1][3].value == "ŠIFRA LOKACIJE",
                    self.__ws_economy[1][4].value == "NAZIV TRANSFORMATORSKE STANICE",
                    self.__ws_economy[1][5].value == "ENPV (HRK)"
                    ]):
            return False
        return True
    # </editor-fold>

    # <editor-fold desc="#####  Check Input Data  #####">
    def check_input_data_line(self, row: int):
        # Object info
        Validate.same_object_info(self.__pwin, self.__ws_condition, self.__ws_significance, self.__ws_economy, row)

        # Condition
        Validate.control_center(self.__pwin, self.__ws_condition, row, 0)
        Validate.if_zero_and_one(self.__pwin, self.__ws_condition, row, 5)
        Validate.year(self.__pwin, self.__ws_condition, row, 6, self.__observed_year)
        Validate.if_between_zero_and_one(self.__pwin, self.__ws_condition, row, 7)
        Validate.if_zero_and_one(self.__pwin, self.__ws_condition, row, 8)
        Validate.if_zero_and_one(self.__pwin, self.__ws_condition, row, 9)

        # Significance
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 5)
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 6)
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 7)
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 8)
        Validate.if_between_zero_and_one(self.__pwin, self.__ws_significance, row, 9)

        # Economy
        Validate.if_number(self.__pwin, self.__ws_economy, row, 5)

        return True

    def check_input_data_trafo_large(self, row: int):
        # Object info
        Validate.same_object_info(self.__pwin, self.__ws_condition, self.__ws_significance, self.__ws_economy, row)

        # Condition
        Validate.control_center(self.__pwin, self.__ws_condition, row, 0)
        Validate.if_zero_and_one(self.__pwin, self.__ws_condition, row, 5)
        Validate.year(self.__pwin, self.__ws_condition, row, 6, self.__observed_year)
        Validate.if_between_zero_and_one(self.__pwin, self.__ws_condition, row, 7)

        # Significance
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 5)
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 6)
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 7)

        # Economy
        Validate.if_number(self.__pwin, self.__ws_economy, row, 5)

        return True

    def check_input_data_trafo_medium(self, row: int):
        # Object info
        Validate.same_object_info(self.__pwin, self.__ws_condition, self.__ws_significance, self.__ws_economy, row)

        # Condition
        Validate.control_center(self.__pwin, self.__ws_condition, row, 0)
        Validate.if_zero_and_one(self.__pwin, self.__ws_condition, row, 5)
        Validate.year(self.__pwin, self.__ws_condition, row, 6, self.__observed_year)
        Validate.if_between_zero_and_one(self.__pwin, self.__ws_condition, row, 7)

        # Significance
        Validate.if_number_positive(self.__pwin, self.__ws_significance, row, 5)
        Validate.if_between_zero_and_one(self.__pwin, self.__ws_significance, row, 6)
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 7)

        # Economy
        # Validate.if_number(self.__pwin, self.__ws_economy, row, 5)

        return True

    def check_input_data_station(self, row: int):
        # Object info
        Validate.same_object_info(self.__pwin, self.__ws_condition, self.__ws_significance, self.__ws_economy, row)

        # Condition
        Validate.control_center(self.__pwin, self.__ws_condition, row, 0)
        Validate.year(self.__pwin, self.__ws_condition, row, 5, self.__observed_year)
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_condition, row, 6)
        Validate.if_zero_and_one(self.__pwin, self.__ws_condition, row, 7)
        Validate.if_zero_and_one(self.__pwin, self.__ws_condition, row, 8)
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_condition, row, 9)

        # Significance
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 5)
        Validate.if_number_positive(self.__pwin, self.__ws_significance, row, 6)
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 7)
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 8)
        Validate.if_zero_and_half_and_one(self.__pwin, self.__ws_significance, row, 9)

        # Economy
        Validate.if_number(self.__pwin, self.__ws_economy, row, 5)

        return True
    # </editor-fold>

    # <editor-fold desc="#####  Read Input Data  #####">
    def read_line(self):
        self.__pwin.logDebug.info("Call:ImportData.read_line()")
        for row in range(2, self.__ws_condition.max_row + 1):
            if self.check_input_data_line(row):
                obj = InvestmentObject(object_type=ObjectType.Line,
                                       row_number=row,
                                       control_center=ControlCenter[self.__ws_condition[row][0].value],
                                       investment_id=self.__ws_condition[row][1].value,
                                       item=self.__ws_condition[row][2].value,
                                       code=self.__ws_condition[row][3].value,
                                       name=self.__ws_condition[row][4].value)

                obj.condition = Condition(technical_condition=self.__ws_condition[row][5].value,
                                          year_of_construction=self.__ws_condition[row][6].value,
                                          unavailability=self.__ws_condition[row][7].value,
                                          examination_and_diagnostics=self.__ws_condition[row][8].value,
                                          other_indicators=self.__ws_condition[row][9].value)

                obj.significance = Significance(transmission_capacity_increase_need=self.__ws_significance[row][5].value,
                                                section_safety=self.__ws_significance[row][6].value,
                                                permanent_unavailability_risk=self.__ws_significance[row][7].value,
                                                damage_evaluation=self.__ws_significance[row][8].value,
                                                revitalization_infulance=self.__ws_significance[row][9].value)

                obj.economy = Economy(enpv=self.__ws_economy[row][5].value)
                self.investmentObjects.append(obj)
        self.__pwin.logDebug.info("Return:ImportData.read_line()")

    def read_transformer_large(self):
        self.__pwin.logDebug.info("Call:ImportData.read_transformer_large()")
        for row in range(2, self.__ws_condition.max_row + 1):
            if self.check_input_data_trafo_large(row):
                obj = InvestmentObject(object_type=ObjectType.TransformerLarge,
                                       row_number=row,
                                       control_center=ControlCenter[self.__ws_condition[row][0].value],
                                       investment_id=self.__ws_condition[row][1].value,
                                       item=self.__ws_condition[row][2].value,
                                       code=self.__ws_condition[row][3].value,
                                       name=self.__ws_condition[row][4].value)

                obj.condition = Condition(examination_and_diagnostics=self.__ws_condition[row][5].value,
                                          year_of_construction=self.__ws_condition[row][6].value,
                                          unavailability=self.__ws_condition[row][7].value)

                obj.significance = Significance(section_safety=self.__ws_significance[row][5].value,
                                                damage_evaluation=self.__ws_significance[row][6].value,
                                                permanent_unavailability_risk=self.__ws_significance[row][7].value)

                obj.economy = Economy(enpv=self.__ws_economy[row][5].value)
                self.investmentObjects.append(obj)
        self.__pwin.logDebug.info("Return:ImportData.read_transformer_large()")

    def read_transformer_medium(self):
        self.__pwin.logDebug.info("Call:ImportData.read_transformer_medium()")
        for row in range(2, self.__ws_condition.max_row + 1):
            if self.check_input_data_trafo_medium(row):
                obj = InvestmentObject(object_type=ObjectType.TransformerMedium,
                                       row_number=row,
                                       control_center=ControlCenter[self.__ws_condition[row][0].value],
                                       investment_id=self.__ws_condition[row][1].value,
                                       item=self.__ws_condition[row][2].value,
                                       code=self.__ws_condition[row][3].value,
                                       name=self.__ws_condition[row][4].value)

                obj.condition = Condition(examination_and_diagnostics=self.__ws_condition[row][5].value,
                                          year_of_construction=self.__ws_condition[row][6].value,
                                          unavailability=self.__ws_condition[row][7].value)

                obj.significance = Significance(transmitted_energy=self.__ws_significance[row][5].value,
                                                max_load_app_power_ratio=self.__ws_significance[row][6].value,
                                                damage_evaluation=self.__ws_significance[row][7].value)

                obj.economy = Economy(enpv=0)
                self.investmentObjects.append(obj)
        self.__pwin.logDebug.info("Return:ImportData.read_transformer_medium()")

    def read_station(self):
        self.__pwin.logDebug.info("Call:ImportData.read_station()")
        for row in range(2, self.__ws_condition.max_row + 1):
            if self.check_input_data_station(row):
                obj = InvestmentObject(object_type=ObjectType.Station,
                                       row_number=row,
                                       control_center=ControlCenter[self.__ws_condition[row][0].value],
                                       investment_id=self.__ws_condition[row][1].value,
                                       item=self.__ws_condition[row][2].value,
                                       code=self.__ws_condition[row][3].value,
                                       name=self.__ws_condition[row][4].value)

                obj.condition = Condition(year_of_construction=self.__ws_condition[row][5].value,
                                          condition_primary_equipment=self.__ws_condition[row][6].value,
                                          condition_secondary_equipment=self.__ws_condition[row][7].value,
                                          condition_auxiliary_equipment=self.__ws_condition[row][8].value,
                                          condition_construction_parts=self.__ws_condition[row][9].value)

                obj.significance = Significance(short_circuit_level=self.__ws_significance[row][5].value,
                                                annual_energy=self.__ws_significance[row][6].value,
                                                downtime_risk=self.__ws_significance[row][7].value,
                                                downtime_damage=self.__ws_significance[row][8].value,
                                                planned_connections_number=self.__ws_significance[row][9].value)

                obj.economy = Economy(enpv=self.__ws_economy[row][5].value)
                self.investmentObjects.append(obj)
        self.__pwin.logDebug.info("Return:ImportData.read_station()")
    # </editor-fold>

    # <editor-fold desc="#####  Create Rang List  #####">
    def update_max_values(self):
        self.__max_age = 0
        self.__max_unavailability = 0
        self.__max_revitalization_infulance = 0
        self.__max_enpv = 0
        self.__max_transmitted_energy = 0
        self.__max_ratio = 0
        self.__max_energy = 0

        for obj in self.investmentObjects:
            if obj.is_in_calculation:
                if self.__observed_year - obj.condition.year > self.__max_age:
                    self.__max_age = self.__observed_year - obj.condition.year
                if obj.condition.unavailability is not None and obj.condition.unavailability > self.__max_unavailability:
                    self.__max_unavailability = obj.condition.unavailability
                if obj.significance.revitalization_infulance is not None and obj.significance.revitalization_infulance > self.__max_unavailability:
                    self.__max_revitalization_infulance = obj.significance.revitalization_infulance
                if obj.economy.enpv > self.__max_enpv:
                    self.__max_enpv = obj.economy.enpv
                if obj.significance.transmitted_energy is not None and obj.significance.transmitted_energy > self.__max_transmitted_energy:
                    self.__max_transmitted_energy = obj.significance.transmitted_energy
                if obj.significance.max_load_app_power_ratio is not None and obj.significance.max_load_app_power_ratio > self.__max_ratio:
                    self.__max_ratio = obj.significance.max_load_app_power_ratio
                if obj.significance.annual_energy is not None and obj.significance.annual_energy > self.__max_energy:
                    self.__max_energy = obj.significance.annual_energy

    def calculate_indexes(self, weight_factors):
        """
        Calculates index for every investment object that is still in calculation
        :param weight_factors: weightFactors.WeightFactors object
        :return: None
        """
        self.update_max_values()
        for obj in self.investmentObjects:
            if obj.is_in_calculation:
                if obj.type == ObjectType.Line:
                    obj.condition.calculate_index_line(weight_factors.data["condition"], self.__observed_year, self.__max_age, self.__max_unavailability)
                    obj.significance.calculate_index_line(weight_factors.data["significance"], self.__max_revitalization_infulance)
                elif obj.type == ObjectType.TransformerLarge:
                    obj.condition.calculate_index_trafo_large(weight_factors.data["condition"], self.__observed_year, self.__max_age, self.__max_unavailability)
                    obj.significance.calculate_index_trafo_large(weight_factors.data["significance"])
                elif obj.type == ObjectType.TransformerMedium:
                    obj.condition.calculate_index_trafo_med(weight_factors.data["condition"], self.__observed_year, self.__max_age, self.__max_unavailability)
                    obj.significance.calculate_index_trafo_med(weight_factors.data["significance"], self.__max_transmitted_energy, self.__max_ratio)
                elif obj.type == ObjectType.Station:
                    obj.condition.calculate_index_station(weight_factors.data["condition"], self.__observed_year, self.__max_age)
                    obj.significance.calculate_index_station(weight_factors.data["significance"], self.__max_energy)
                obj.economy.calculate_index([1], self.__max_enpv)
                obj.calculate_index(weight_factors.data["all"])

    def return_list_sort_by_index_desc(self):
        rank_list = [obj for obj in self.investmentObjects if obj.is_in_calculation]
        rank_list.sort(key=lambda obj: obj.index, reverse=True)
        print("---Iteration----------------------------------")
        for x in rank_list:
            print(x.name, x.index)
        print("----------------------------------------------")
        return rank_list
    # </editor-fold>
